from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, ListView
from django.views.generic.edit import UpdateView, DeleteView, CreateView
from django.views.generic.detail import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404
from django.views import View
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from .models import *
from .forms import *
from django.core.mail import BadHeaderError, send_mail
from django.http import HttpResponse
from django.contrib.auth.backends import UserModel
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth.views import PasswordChangeView
from django.db.models import Q
from django.contrib.auth.forms import PasswordResetForm
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.conf import settings
import pandas as pd
import json
import openpyxl
import tabula
from django.utils.translation import get_language, activate, gettext
from collections import namedtuple
from django.db.models import Q


def admin_home(request):
    return render(request, 'admin_templates/index.html')

@csrf_exempt
def authorization(request):
    success_url = reverse_lazy('admin_panel')
    username = request.POST['username']
    password = request.POST['password']

    try:
        account = authenticate(username=CustomUser.objects.get(email=username).username, password=password)
        

        if account is not None and request.method == 'POST':
            login(request, account)
            return redirect(success_url)
        else:
            messages.error(request, "Неверный логин или пароль!")
            return redirect('login')
    except:
        if username == "":
            messages.error(request, "Введите Username или E-mail")
            return redirect('login')
        elif password == "":
            messages.error(request, "Введите пароль")
            return redirect('login')
        else:
            account = authenticate(username=username, password=password)
            if account is not None and request.method == 'POST':
                login(request, account)
                return redirect(success_url)
            else:
                messages.error(request, "Неверный логин или пароль!")
                return redirect('login')


def logout_page(request):
    logout(request)
    return redirect('login')

def translate(language):
    cur_language = get_language()
    try:
        activate(language)
    finally:
        activate(cur_language)
    return cur_language

# Admin-Panel Views
class AdminMain(LoginRequiredMixin, ListView):
    login_url = 'doLogin' #указывает URL-адрес, по которому будет перенаправлен пользователь, если он не авторизовался.
    model = CustomUser #указывает модель ( CustomUser), на которой работает представление.
    paginate_by = 5#указывает количество объектов, отображаемых на странице.
    template_name = "admin_templates/admin.html"#указывает шаблон, используемый для визуализации представления.

    #extra_context указывает дополнительные данные контекста, которые должны быть переданы в шаблон.
    extra_context = {
           "is_active" : "main-panel",
          # "all_entries" : count,
    } 

def password_reset_request(request):
    # Установить учетные данные электронной почты
    settings.EMAIL_HOST_USER='csmbishkek@gmail.com' 
    settings.EMAIL_HOST_PASSWORD='hswhmrcllyxoqbvx' 
    if request.method == "POST": # указывает на проверку, что метод запроса должен быть ПОСТ
        # Создайте экземпляр формы UserPasswordResetForm с отправленными данными
        password_reset_form = UserPasswordResetForm(request.POST) 
        if password_reset_form.is_valid():
            # Получить адрес электронной почты, введенный в форму
            data = password_reset_form.cleaned_data['email']
            print(data)
            # Проверить, есть ли связанные пользователи с введенным адресом электронной почты
            associated_users = CustomUser.objects.filter(Q(email=data))
            if associated_users.exists():
                for user in associated_users:
                    # Подготовьте параметры электронной почты
                    subject = "Запрос на сброс данных"
                    email_template_name = "admin_templates/pages/user/password_reset_email.html"
                    c = {
                        "email":user.email,
                        'domain':'http://127.0.0.1:8000',
                        'site_name': 'Website',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                        }
                    email = render_to_string(email_template_name, c)
                    
                    try:
                        # Отправьте электронное письмо для сброса пароля
                        send_mail(subject, email, 'vkrsupp@gmail.com' , [user.email], fail_silently=False)
                    except BadHeaderError:
                        
                        return HttpResponse('Invalid header found.')
                    # Перенаправление на страницу сброса пароля
                    return redirect ("/accounts/password_reset/done/")
    # Создать новый экземпляр формы UserPasswordResetForm
    password_reset_form = UserPasswordResetForm()
    #Перейти на страницу сброса пароля с формой
    return render(request=request, template_name="admin_templates/pages/user/password_reset.html", context={"password_reset_form":password_reset_form})

class ProfileView(View):
    model = User
    form_class = UpdateUserForm
    success_url = reverse_lazy('update_profile')
    

class ProfileUpdateView(LoginRequiredMixin, SuccessMessageMixin, ProfileView, UpdateView):
    login_url = "login_page"
    success_message = "Данные успешно изменены"
    template_name = 'admin_templates/pages/user/update.html'
    def get_object(self, queryset=None):
        '''This method will load the object
           that will be used to load the form
           that will be edited'''
        return self.request.user
    
class ChangePasswordView(SuccessMessageMixin, PasswordChangeView):
    template_name = 'admin_templates/pages/user/change_password.html'
    success_message = "Пароль успешно изменён"
    success_url = reverse_lazy('update_profile')



def det_category_id(request, region):
    cat = Categories.objects.get(Q(Name_of_category=request) & Q(Region=region))

    return cat

def Conver_PDF_to_Excel(request):
    
    return render(request, 'admin_templates/pages/convert_to_excel.html') 


def edititems(item): 
    for i in range(len(item)): 
        value = item[i] 
        if value == "-": 
            item[i] = '0,0' 
        elif "±" in str(value): 
            item[i] = value.split("±")[0] 
        elif isinstance(value, float) or isinstance(value, int): 
            item[i] = str(value).replace('.', ',') 
     
    return item


# IMPORT DATA
def Import_Excel_pandas(request):
    if request.method =="POST":
        print("Okay")
        lang = request.POST.get('lang')
        excelfile = request.FILES['myfile']
        workbook = openpyxl.load_workbook(excelfile, data_only=True)
        column_index = 9
        column_index_for_category = 1
        column_index_for_ingredients = 2
        column_data_category = []
        column_data_ingredient = []
        column_data_chemicalingredients = []
        column_data_aminoacidingredients = []
        column_data_mineralingredients= []
        column_data_fatacidsingredients = []
        regions = workbook.sheetnames
        message_from_regions = import_data_Regions(regions,lang)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for cell in sheet.iter_rows(min_col=column_index_for_category, max_col=column_index_for_category,values_only=True):
                column_data_category.append(cell[0])

            first_row_values = []
            for col_num, cell in enumerate(sheet[1], 1):
                first_row_values.append({'value': cell.value, 'column_number': col_num})
            # Определите интересующие вас заголовки
            target_headers = {'Аминокислотый состав', 'Минеральный состав', 'Жирно-кислотный состав'}

            # Получите все значения и номера столбцов в первой строке для интересующих вас заголовков
            result_list = []
            for col_num, cell in enumerate(sheet[1], 1):
                if cell.value in target_headers:
                    result_list.append({'column_number': col_num, 'value': cell.value})

            message_from_category = import_data_Categories(sheet.title, column_data_category,lang)
            
            if message_from_category == "Import success!":
                # Взять 2 столбец для добавления в ингредиент и для последующиъ действий
                for cell in sheet.iter_rows(min_col=column_index_for_ingredients, max_col=column_index_for_ingredients,values_only=True):
                    column_data_ingredient.append(cell[0])
                
                message_from_products = import_data_Ingredients(sheet.title,column_data_category, column_data_ingredient,lang)
                
                if message_from_products == "Import success!":
                    #Взять следующие столбцы для добавления химического состава
                    for row in sheet.iter_rows(min_col = 3,max_col=column_index, values_only=True):
                        row_data = list(row)
                        column_data_chemicalingredients.append(row_data)
                    message_from_chemicalingredients = import_data_ChemicalsIngredients(sheet.title,column_data_category, column_data_ingredient,column_data_chemicalingredients)

                    for row in sheet.iter_rows(min_col = result_list[0]['column_number']+1,max_col=result_list[1]['column_number']-1, values_only=True):
                        row_data = list(row)
                        column_data_aminoacidingredients.append(row_data)
                    message_from_aminoiacidsngredients = import_data_AminoAcidsIngredients(sheet.title,column_data_category, column_data_ingredient,column_data_aminoacidingredients)

                    for row in sheet.iter_rows(min_col = result_list[1]['column_number']+1,max_col=result_list[2]['column_number']-1, values_only=True):
                        row_data = list(row)
                        column_data_mineralingredients.append(row_data)
                    message_from_mineralingredients = import_data_MineralIngredients(sheet.title,column_data_category, column_data_ingredient,column_data_mineralingredients)

                    for row in sheet.iter_rows(min_col = result_list[2]['column_number']+1,max_col=len(first_row_values)-1, values_only=True):
                        row_data = list(row)
                        column_data_fatacidsingredients.append(row_data)
                    message_from_fatcidsingredients = import_data_FatAcidsIngresients(sheet.title,column_data_category, column_data_ingredient,column_data_fatacidsingredients,lang)
            # Ваши действия с полученными данными
            # print(f"Categories ---> {column_data_category}")
            # print(f"Ingredients ---> {column_data_ingredient}")
            # print(f"Data ---> {column_data_chemicalingredients}")
            column_data_category.clear()
            column_data_ingredient.clear()
            column_data_chemicalingredients.clear()
            column_data_aminoacidingredients.clear()
            column_data_mineralingredients.clear()
    else:
        print("Not Okay")
    return render(request, 'admin_templates/pages/Import_excel_db.html')

def import_data_Regions(regions,lang):
    try:
        for item in regions:
            if not Regions.objects.filter(region=item, language=lang).exists():
                region = Regions(region=item, language=lang)
                region.save()
        message = "Import success!"
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred: {str(e)}")
    return message

def import_data_Categories(sheet,data,lang):
    # print("Это запись данных в категории")
    # print(f"Что получил---------------------")
    # print(f"Регион: {sheet}")
    # print(f"Данные: {data}")
    try:
        for item in data[1:]:
            region_obj = Regions.objects.filter(region=sheet).first()
            if not Categories.objects.filter(Name_of_category=item, Region=region_obj, language=lang).exists():
                if item is not None:
                    category = Categories(Name_of_category=item, Region=region_obj,language=lang)
                    #print(f"Что записывает: {category}")
                    category.save()
        message = "Import success!"
    except Exception as e:
        message = "Import not success!" 
        print(f"An error occurred 1: {str(e)}")
    return message

def import_data_Ingredients(sheet,category,data,lang):
    # print("Это запись данных в ингредиенты")
    # print(f"Что получил---------------------")
    # print(f"Регион: {sheet}")
    # print(f"Категория: {category}")
    # print(f"Данные: {data}")
    try:
        for item, itemcat in zip(data[1:], category[1:]):
            #print(f"Лист категорий: {itemcat}")
            region_obj = Regions.objects.filter(region=sheet).first()
            category_obj = Categories.objects.filter(Name_of_category=itemcat, Region=region_obj).first()
            #print(f"Что получил по категориям: {category_obj}")
            if not Products.objects.filter(attribute_name=item, Category=category_obj, language=lang).exists():
                if item is not None:
                    product = Products(attribute_name=item, Category=category_obj,language=lang)
                    #print(f"Что записывает: {product}")
                    product.save()
        message = "Import success!"
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Ingredients: {str(e)}")
    return message

def import_data_ChemicalsIngredients(sheet,category,ingredients, data):
    # print("Это запись данных в Химический состав")
    # print(f"Что получил---------------------")
    # print(f"Регион: {sheet}")
    # print(f"Категории: {category}")
    # print(f"Индгредиенты: {ingredients}")
    # print(f"Данные: {data}")
    try:
        for item, itemcat, itemingr in zip(data[1:], category[1:], ingredients[1:]):
            #print(f"Что получает: {item}")
            item = edititems(item)
            region_obj = Regions.objects.filter(region=sheet).first()
            category_obj = Categories.objects.filter(Name_of_category=itemcat, Region=region_obj).first()
            product_obj = Products.objects.filter(attribute_name=itemingr, Category=category_obj).first()
            
            if item[0] is not None:
                if not Chemicals.objects.filter(product=product_obj).exists():
                    if item[6] is not None:
                        chemical = Chemicals(
                            product=product_obj,
                            soluable_solids=float(item[0].replace(',', '.')),
                            ascorbic_acids=float(item[1].replace(',', '.')),
                            ash_content=float(item[2].replace(',', '.')),
                            moisture=float(item[3].replace(',', '.')),
                            protein=float(item[4].replace(',', '.')),
                            fat=float(item[5].replace(',', '.')),
                            carbohydrates=float(item[6].replace(',', '.')),
                        )
                        chemical.save()
                    else:
                        chemical = Chemicals(
                            product=product_obj,
                            soluable_solids=float(item[0].replace(',', '.')),
                            ascorbic_acids=float(item[1].replace(',', '.')),
                            ash_content=float(item[2].replace(',', '.')),
                            moisture=float(item[3].replace(',', '.')),
                            protein=float(item[4].replace(',', '.')),
                            fat=float(item[5].replace(',', '.'))
                        )
                        chemical.save()

        message = "Import success!"
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Chemical ing: {str(e)}")
    return message

def import_data_AminoAcidsIngredients(sheet,category,ingredients, data):
    # print("Это запись данных в Химический состав")
    # print(f"Что получил---------------------")
    # print(f"Регион: {sheet}")
    # print(f"Категории: {category}")
    # print(f"Индгредиенты: {ingredients}")
    # print(f"Данные: {data}")
    try:
        for item, itemcat, itemingr in zip(data[1:], category[1:], ingredients[1:]):
            #print(f"Что получает: {item}")
            item = edititems(item)
            region_obj = Regions.objects.filter(region=sheet).first()
            category_obj = Categories.objects.filter(Name_of_category=itemcat, Region=region_obj).first()
            product_obj = Products.objects.filter(attribute_name=itemingr, Category=category_obj).first()
            if item[0] is not None:
                if not AminoAcidComposition.objects.filter(product=product_obj).exists():
                    aminoacids = AminoAcidComposition(
                        product=product_obj,
                        asparing=float(item[0].replace(',', '.')),
                        glutamin=float(item[1].replace(',', '.')),
                        serin=float(item[2].replace(',', '.')),
                        gistidin=float(item[3].replace(',', '.')),
                        glitsin=float(item[4].replace(',', '.')),
                        treonin=float(item[5].replace(',', '.')),
                        arginin=float(item[6].replace(',', '.')),
                        alanin=float(item[7].replace(',', '.')),
                        tirosin=float(item[8].replace(',', '.')),
                        tsistein=float(item[9].replace(',', '.')),
                        valin=float(item[10].replace(',', '.')),
                        metionin=float(item[11].replace(',', '.')),
                        triptofan=float(item[12].replace(',', '.')),
                        fenilalalin=float(item[13].replace(',', '.')),
                        izoleitsin=float(item[14].replace(',', '.')),
                        leitsin=float(item[15].replace(',', '.')),
                        lisin=float(item[16].replace(',', '.')),
                        prolin=float(item[17].replace(',', '.'))
                    )
                    aminoacids.save()
        message = "Import success!"
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Chemical: {str(e)}")
    return message

def import_data_MineralIngredients(sheet, category, ingredients, data):
    try:
        for item, itemcat, itemingr in zip(data[1:], category[1:], ingredients[1:]):
            
            item = edititems(item)
            region_obj = Regions.objects.filter(region=sheet).first()
            category_obj = Categories.objects.get(Name_of_category=itemcat, Region=region_obj)
            product_obj = Products.objects.get(attribute_name=itemingr, Category=category_obj)
            if item[0] is not None:
                if not MineralComposition.objects.filter(product=product_obj).exists():
                    
                    mineral = MineralComposition(
                        product=product_obj,
                        Ca=float(item[0].replace(',', '.')),
                        Na=float(item[1].replace(',', '.')),
                        K=float(item[2].replace(',', '.')),
                        P=float(item[3].replace(',', '.')),
                        Mn=float(item[4].replace(',', '.')),
                        Zn=float(item[5].replace(',', '.')),
                        Se=float(item[6].replace(',', '.')),
                        Cu=float(item[7].replace(',', '.')),
                        Fe=float(item[8].replace(',', '.')),
                        I=float(item[9].replace(',', '.')),
                        B=float(item[10].replace(',', '.')),
                        Li=float(item[11].replace(',', '.')),
                        Al=float(item[12].replace(',', '.')),
                        Mg=float(item[13].replace(',', '.')),
                        V=float(item[14].replace(',', '.')),
                        Ni=float(item[15].replace(',', '.')),
                        Co=float(item[16].replace(',', '.')),
                        Cr=float(item[17].replace(',', '.')),
                        Sn=float(item[18].replace(',', '.'))
                    )
                    mineral.save()
        message = "Import success!"
   
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Mineral ing: {str(e)}")
    return message

def types_acid(request, lang):
    if request == "Насыщенные жирные кислоты, %" and lang == "Русский":
        request = "Насыщенные жирные кислоты, %"
        return request
    elif request == "Мононенасыщенные жирные кислоты, %" and lang == "Русский":
        request = "Мононенасыщенные жирные кислоты, %"
        return request
    elif request == "Полиненасыщенные жирные кислоты, %" and lang == "Русский":
        request = "Полиненасыщенные жирные кислоты, %"
        return request
    elif request == "Saturated fatty acids, %" and lang == "English":
        request = "Saturated fatty acids, %"
        return request
    elif request == "Monounsaturated fatty acids, %" and lang == "English":
        request = "Monounsaturated fatty acids, %"
        return request
    elif request == "Polyunsaturated fatty acids, %" and lang == "English":
        request = "Polyunsaturated fatty acids, %"
        return request
    else:
        return request

def import_data_FatAcidsIngresients(sheet,category,ingredients,data,lang):
    try:
        item_for_acid = data[0]
        fat_acids = import_data_fatacids(data[0], lang)
        fat_acids_type = import_data_fatacidstype(data[0],lang)
        
        for itemcat, itemingr, item in zip(category[1:], ingredients[1:], data[1:]):
            item = edititems(item)
            i=0
            fat_acid_type = 'Насыщенные жирные кислоты, %'
            category_obj = Categories.objects.get(Name_of_category=itemcat, Region=Regions.objects.filter(region=sheet).first())
            product_obj = Products.objects.get(attribute_name=itemingr, Category=category_obj)
                
            for k, fatacid in zip(item, item_for_acid):
                fat_acid_type_value = types_acid(fatacid, lang)
                
                if fat_acid_type_value == 'Мононенасыщенные жирные кислоты, %':
                    fat_acid_type = 'Мононенасыщенные жирные кислоты, %'
                elif fat_acid_type_value == 'Monounsaturated fatty acids, %':
                    fat_acid_type = 'Monounsaturated fatty acids, %'
                if fat_acid_type_value == 'Полиненасыщенные жирные кислоты, %':
                    fat_acid_type = 'Полиненасыщенные жирные кислоты, %'
                elif fat_acid_type_value =='Polyunsaturated fatty acids, %':
                    fat_acid_type = 'Polyunsaturated fatty acids, %'
                elif fat_acid_type_value == 'Saturated fatty acids, %':
                    fat_acid_type = 'Saturated fatty acids, %'
                
                types_fataacid = FatAcidsType.objects.get(name=fat_acid_type, language = lang)
                if k is not None:
                    object_fat_acid = FatAcids.objects.filter(name=fatacid).first()
                    fatacids, created = FatAcidCompositionOfMeal.objects.get_or_create(
                        product=product_obj,
                        types=types_fataacid,
                        fat_acid=object_fat_acid,
                        language=lang,
                        defaults={'value': float(k.replace(',', '.'))}
                    )

                    if not created:
                        fatacids.save()
        message = "Import success!"
    except Exception as e:
        message = "Import not success!"
        print(f"I: {str(i)}")
        print(f"An error occurred FatAcids ing: {str(e)}")
    return message

def import_data_fatacids(item, lang):
    try:
        for value in item:
            if item[0] is not None:
                if not FatAcids.objects.filter(name=value, language=lang).exists():
                    fatacids = FatAcids(name=value, language=lang)
                    fatacids.save()
        return item
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Fat Acids: {str(e)}")
    return message

def import_data_fatacidstype(item,lang):
    try:
        items = []
        for value in item:
            value = value.split('*')
            if value[0] == "Насыщенные жирные кислоты, %" or value[0]== "Saturated fatty acids, %":
                if item[0] is not None:
                    if not FatAcidsType.objects.filter(name=value[0],language=lang).exists():
                        fatacidstype = FatAcidsType(name=value[0],language=lang)
                        fatacidstype.save()
                        items.append(fatacidstype)
            elif value[0] =="Мононенасыщенные жирные кислоты, %" or value[0]== "Monounsaturated fatty acids, %":
                if item[0] is not None:
                    if not FatAcidsType.objects.filter(name=value[0],language=lang).exists():
                        fatacidstype = FatAcidsType(name=value[0],language=lang)
                        fatacidstype.save()
                        items.append(fatacidstype)
            elif value[0] =="Полиненасыщенные жирные кислоты, %" or value[0]== "Polyunsaturated fatty acids, %":
                if item[0] is not None:
                    if not FatAcidsType.objects.filter(name=value[0],language=lang).exists():
                        fatacidstype = FatAcidsType(name=value[0],language=lang)
                        fatacidstype.save()
                        items.append(fatacidstype)
        return items
    except Exception as e:
        message = "Import not success!"
        print(f"An error occurred Fat Acids Type: {str(e)}")
    return message

@csrf_exempt
def check_email_exist(request):
    email = request.POST.get("email")
    user_obj = CustomUser.objects.filter(email=email).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


@csrf_exempt
def check_username_exist(request):
    username = request.POST.get("username")
    user_obj = CustomUser.objects.filter(username=username).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)

def add_staff(request):
    return render(request, "admin_templates/pages/add_staff.html")


def add_staff_save(request):

    if request.method != "POST":
        messages.error(request, "Неверный метод ")
        return redirect('add_staff')
    
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        user_type = request.POST.get('user_type')

        try:
            user = CustomUser.objects.create_user(username=username, password=password, email=email, first_name=first_name, last_name=last_name, user_type=user_type)
            user.save()
            
            messages.success(request, "Регистрация сотрудника прошла успешно!")
            return redirect('add_staff')
        
        except:
            if username == "":
                messages.error(request, "Введите Username или E-mail")
                return redirect('add_staff')
            elif first_name == "":
                messages.error(request, "Введите Имя")
                return redirect('add_staff')
            elif last_name == "":
                messages.error(request, "Введите Фамилию")
                return redirect('add_staff')
            elif password == "":
                messages.error(request, "Введите пароль")
                return redirect('add_staff')
            elif user_type == "" or user_type == None:
                messages.error(request, "Выберите тип пользователя!")
                return redirect('add_staff')
            
            return redirect('add_staff')



def manage_staff(request):
    staffs = CustomUser.objects.filter(Q(user_type=2) | Q(user_type=3))

    context = {
        "staffs": staffs
    }
    return render(request, "admin_templates/pages/manage_staff_template.html", context)


def edit_staff(request, staff_id):
    staff = CustomUser.objects.get(id=staff_id)
    user_type_choices = CustomUser.user_type_data

    context = {
        "staff": staff,
        "id": staff_id, 
        "user_type_choices":user_type_choices,
    }

    return render(request, "admin_templates/pages/edit_staff_template.html", context)


def edit_staff_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id = request.POST.get('staff_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        address = request.POST.get('address')
        user_type = request.POST.get('user_type')

        try:
            # INSERTING into Customuser Model
            user = CustomUser.objects.get(id=staff_id)
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.username = username
            user.user_type = user_type
            user.save()
            
            # INSERTING into Staff Model
            staff_model = Staffs.objects.get(admin=staff_id)
            staff_model.address = address
            staff_model.save()

            messages.success(request, "Обновление записи выполнено успешно.")
            return redirect('/admin-panel/edit_staff/'+staff_id)

        except:
            messages.error(request, "Возникли ошибки при обновлении записи.")
            return redirect('/admin-panel/edit_staff/'+staff_id)



def delete_staff(request, staff_id):
    staff = CustomUser.objects.get(id=staff_id)
    try:
        staff.delete()
        messages.success(request, "Удаление записи выполнено успешно.")
        return redirect('manage_staff')
    except:
        messages.error(request, "Возникли ошибки при удалении записи.")
        return redirect('manage_staff')

#Regions CRUD operations   
class RegionsView(View):
    model = Regions
    form_class = RegionsForm
    active_panel = "regions-panel"
    login_url = "login_page"
    success_url = reverse_lazy("regions_create")
    extra_context = {
        "is_active" : active_panel,
        "active_regions" : "active",
        "expand_regions" : "show",
        }
    
class RegionsListView(LoginRequiredMixin, RegionsView, ListView):
    template_name = "admin_templates/pages/regions/regions_list.html"
    paginate_by = 10

class RegionsCreateView(LoginRequiredMixin, SuccessMessageMixin, RegionsView, CreateView):
    template_name = 'admin_templates/pages/regions/regions_form.html'
    success_message = "Запись успешно Добавлена!"
    
class RegionsUpdateView(LoginRequiredMixin, SuccessMessageMixin, RegionsView, UpdateView):
    template_name = "admin_templates/pages/regions/regions_form.html"
    success_url = reverse_lazy("regions_all")
    success_message = "Запись успешно Обновлена!"

class RegionsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            Regions.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('regions_all')

def regions_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = Regions.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("regions_all")

        except Regions.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("regions_delete")

    context={
        "messages":messages,
    }

    return render(request, "admin_templates/pages/regions/regions_all_delete_confirm.html", context)
   
def regions_delete(request, id):
    context = {
            "is_active" : "regions-panel",
            "active_regions" : "active",
            "expand_regions" : "show",
    }
    obj = get_object_or_404(Regions, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("regions_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("regions_delete")

    return render(request, "admin_templates/pages/regions/regions_delete_confirm.html", context)  


# Products CRUD operations

class ProductsView(View):
    model = Products
    form_class = ProductsForm
    active_panel = "products-panel"
    login_url = "login_page"
    success_url = reverse_lazy("products_create")
    extra_context = {
        "is_active" : active_panel,
        "active_products" : "active",
        "expand_products" : "show",
        }
    
class ProductsListView(LoginRequiredMixin, ProductsView, ListView):
    template_name = "admin_templates/pages/products/products_list.html"
    paginate_by = 10

class ProductsCreateView(LoginRequiredMixin, SuccessMessageMixin, ProductsView, CreateView):
    template_name = 'admin_templates/pages/products/products_form.html'
    success_message = "Запись успешно Добавлена!"  

class ProductsUpdateView(LoginRequiredMixin, SuccessMessageMixin, ProductsView, UpdateView):
    template_name = "admin_templates/pages/products/products_form.html"
    success_url = reverse_lazy("products_all")
    success_message = "Запись успешно Обновлена!"

class ProductsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            Products.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('products_all')
    

def products_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = Products.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("products_all")

        except Products.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("products_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/products/products_all_delete_confirm.html", context)

def products_delete(request, id):
    context = {
            "is_active" : "products-panel",
            "active_products" : "active",
            "expand_products" : "show",
    }
    obj = get_object_or_404(Products, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("products_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("products_delete")
 
    return render(request, "admin_templates/pages/products/products_delete_confirm.html", context)  


# Categories CRUD operations

class CategoriesView(View):
    model = Categories
    form_class = CategoriesForm
    active_panel = "categories-panel"
    login_url = "login_page"
    success_url = reverse_lazy("categories_create")
    extra_context = {
        "is_active" : active_panel,
        "active_categories" : "active",
        "expand_categories" : "show",
        }
    
class CategoriesListView(LoginRequiredMixin, CategoriesView, ListView):
    template_name = "admin_templates/pages/categories/categories_list.html"
    paginate_by = 10

class CategoriesCreateView(LoginRequiredMixin, SuccessMessageMixin, CategoriesView, CreateView):
    template_name = 'admin_templates/pages/categories/categories_form.html'
    success_message = "Запись успешно Добавлена!"  

class CategoriesUpdateView(LoginRequiredMixin, SuccessMessageMixin, CategoriesView, UpdateView):
    template_name = "admin_templates/pages/categories/categories_form.html"
    success_url = reverse_lazy("categories_all")
    success_message = "Запись успешно Обновлена!"

class CategoriesDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            Categories.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('categories_all')

def categories_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = Categories.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("categories_all")

        except Categories.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("categories_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/categories/categories_all_delete_confirm.html", context)

def categories_delete(request, id):
    context = {
            "is_active" : "categories-panel",
            "active_categories" : "active",
            "expand_categories" : "show",
    }
    obj = get_object_or_404(Categories, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("categories_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("categories_delete")
 
    return render(request, "admin_templates/pages/categories/categories_delete_confirm.html", context)


# Typs CRUD operations


# FatAcids CRUD operations

class FatAcidsView(View):
    model = FatAcids
    form_class = FatAcidsForm
    active_panel = "fatacids-panel"
    login_url = "login_page"
    success_url = reverse_lazy("fatacids_create")
    extra_context = {
        "is_active" : active_panel,
        "active_fatacids" : "active",
        "expand_fatacids" : "show",
        }
    
class FatAcidsListView(LoginRequiredMixin, FatAcidsView, ListView):
    template_name = "admin_templates/pages/fatacids/fatacids_list.html"
    paginate_by = 10

class FatAcidsCreateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidsView, CreateView):
    login_url = 'login_page'
    template_name = 'admin_templates/pages/fatacids/fatacids_form.html'
    success_message = "Запись успешно Добавлена!"  

class FatAcidsUpdateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidsView, UpdateView):
    template_name = "admin_templates/pages/fatacids/fatacids_form.html"
    success_url = reverse_lazy("fatacids_create")
    success_message = "Запись успешно Обновлена!"

class FatAcidsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            FatAcids.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('fatacids_all')

def fatacids_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = FatAcids.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("fatacids_all")

        except FatAcids.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("fatacids_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/fatacids/fatacids_all_delete_confirm.html", context)

def fatacids_delete(request, id):
    context = {
            "is_active" : "fatacids-panel",
            "active_fatacids" : "active",
            "expand_fatacids" : "show",
    }
    obj = get_object_or_404(FatAcids, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("fatacids_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("fatacids_delete")
 
    return render(request, "admin_templates/pages/fatacids/fatacids_delete_confirm.html", context) 


#FatAcidsTypes
class FatAcidTypesView(View):
    model = FatAcidsType
    form_class = FatAcidsTypeForm
    active_panel = "fatacidstype-panel"
    login_url = "login_page"
    success_url = reverse_lazy("fatacidstype_create")
    extra_context = {
        "is_active" : active_panel,
        "active_fatacidstype" : "active",
        "expand_fatacidstype" : "show",
        }
    
class FatAcidTypesListView(LoginRequiredMixin, FatAcidTypesView, ListView):
    template_name = "admin_templates/pages/fatacidstype/fatacidstype_list.html"
    paginate_by = 10

class FatAcidTypesCreateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidTypesView, CreateView):
    login_url = 'login_page'
    template_name = 'admin_templates/pages/fatacidstype/fatacidstype_form.html'
    success_message = "Запись успешно Добавлена!"  

class FatAcidTypesUpdateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidTypesView, UpdateView):
    template_name = "admin_templates/pages/fatacidstype/fatacidstype_form.html"
    success_url = reverse_lazy("fatacidstype_create")
    success_message = "Запись успешно Обновлена!"

class FatAcidTypesDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        print(f"selected_ids ---> {selected_ids}")
        try:
            if selected_ids:
                FatAcidsType.objects.filter(id__in=selected_ids).delete()

                messages.success(request, 'Выбранные записи успешно удалены.')
            else:
                messages.error(request, 'Не выбраны записи для удаления.')
        except ValueError:
            messages.error(request, 'Некорректные идентификаторы для удаления.')

        return redirect('fatacidstype_all')

def fatacidstypes_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = FatAcidsType.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("fatacidstype_all")

        except FatAcidsType.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("fatacidstype_delete")

    context={
        "messages":messages,
    }

    return render(request, "admin_templates/pages/fatacidstype/fatacidstype_all_delete_confirm.html", context)

def fatacidstype_delete(request, id):
    context = {
            "is_active" : "fatacidstype-panel",
            "active_fatacidstype" : "active",
            "expand_fatacidstype" : "show",
    }
    obj = get_object_or_404(FatAcidsType, id = id)
    if request.method =="POST":
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("fatacidstype_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("fatacidstype_delete")
 
    return render(request, "admin_templates/pages/fatacidstype/fatacidstype_delete_confirm.html", context)



#FatAcidComposition
class FatAcidCompositionView(View):
    model = FatAcidCompositionOfMeal
    form_class = FatAcidCompositionForm
    active_panel = "fatacidcomposition-panel"
    login_url = "login_page"
    success_url = reverse_lazy("fatacidcomposition_create")
    extra_context = {
        "is_active" : active_panel,
        "active_fatacidcomposition" : "active",
        "expand_fatacidcomposition" : "show",
        }
    
class FatAcidCompositionListView(LoginRequiredMixin, FatAcidCompositionView, ListView):
    template_name = "admin_templates/pages/fatacidcomposition/fatacidcomposition_list.html"
    paginate_by = 10

class FatAcidCompositionCreateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidCompositionView, CreateView):
    login_url = 'login_page'
    template_name = 'admin_templates/pages/fatacidcomposition/fatacidcomposition_form.html'
    success_message = "Запись успешно Добавлена!"  

class FatAcidCompositionUpdateView(LoginRequiredMixin, SuccessMessageMixin, FatAcidCompositionView, UpdateView):
    template_name = "admin_templates/pages/fatacidcomposition/fatacidcomposition_form.html"
    success_url = reverse_lazy("fatacidcomposition_create")
    success_message = "Запись успешно Обновлена!"

class FatAcidCompositionDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        print(f"selected_ids ----->  {selected_ids}")
        try:
            FatAcidCompositionOfMeal.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Выбранные записи успешно удалены.')
            return redirect('fatacidcomposition_all', {"messages":messages})
        except ValueError:
            messages.error(request, 'Некорректные идентификаторы для удаления.')
            return redirect('fatacidcomposition_all', {"messages":messages})


def fatacidcomposition_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = FatAcidCompositionOfMeal.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("fatacidcomposition_all")

        except FatAcidCompositionOfMeal.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("fatacidcomposition_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/fatacidcomposition/fatacidcomposition_all_delete_confirm.html", context)

def fatacidcomposition_delete(request, id):
    context = {
            "is_active" : "fatacidcomposition-panel",
            "active_fatacidcomposition" : "active",
            "expand_fatacidcomposition" : "show",
    }
    obj = get_object_or_404(FatAcidCompositionOfMeal, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("fatacidcomposition_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("fatacidcomposition_delete")
 
    return render(request, "admin_templates/pages/fatacidcomposition/fatacidcomposition_delete_confirm.html", context)



# Mineral compositions CRUD operations

class MineralsView(View):
    model = MineralComposition
    form_class = MineralForm
    active_panel = "minerals-panel"
    success_url = reverse_lazy("minerals_create")
    extra_context = {
        "is_active" : active_panel,
        "active_minerals" : "active",
        "expand_minerals" : "show",
        }
    
class MineralsListView(LoginRequiredMixin, MineralsView, ListView):
    login_url = "login_page"
    template_name = "admin_templates/pages/minerals/minerals_list.html"
    paginate_by = 10

class MineralsCreateView(LoginRequiredMixin, SuccessMessageMixin, MineralsView, CreateView):
    login_url = 'login_page'
    template_name = 'admin_templates/pages/minerals/minerals_form.html'
    success_message = "Запись успешно Добавлена!" 

class MineralsUpdateView(LoginRequiredMixin, SuccessMessageMixin, MineralsView, UpdateView):
    login_url = "login_page"
    template_name = "admin_templates/pages/minerals/minerals_form.html"
    success_url = reverse_lazy("minerals_all")
    success_message = "Запись успешно Обновлена!"

class MineralsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            MineralComposition.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('minerals_all')

def minerals_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = MineralComposition.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("minerals_all")

        except MineralComposition.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("minerals_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/minerals/minerals_all_delete_confirm.html", context)    

def minerals_delete(request, id):
    context = {
            "is_active" : "minerals-panel",
            "active_minerals" : "active",
            "expand_minerals" : "show",
    }
    obj = get_object_or_404(MineralComposition, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("minerals_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("minerals_delete")
 
    return render(request, "admin_templates/pages/minerals/minerals_delete_confirm.html", context) 

# AminoAcids CRUD operations

class AminoAcidsView(View):
    model = AminoAcidComposition
    form_class = AminoAcidsForm
    login_url = "login_page"
    active_panel = "aminoacids-panel"
    success_url = reverse_lazy("aminoacids_create")
    extra_context = {
        "is_active" : active_panel,
        "active_aminoacids" : "active",
        "expand_aminoacids" : "show",
        }
    
class AminoAcidsListView(LoginRequiredMixin, AminoAcidsView, ListView):
    template_name = "admin_templates/pages/aminoacids/aminoacids_list.html"
    paginate_by = 10

class AminoAcidsCreateView(LoginRequiredMixin, SuccessMessageMixin, AminoAcidsView, CreateView):
    template_name = 'admin_templates/pages/aminoacids/aminoacids_form.html'
    success_message = "Запись успешно Добавлена!" 

class AminoAcidsUpdateView(LoginRequiredMixin, SuccessMessageMixin, AminoAcidsView, UpdateView):
    template_name = "admin_templates/pages/aminoacids/aminoacids_form.html"
    success_message = "Запись успешно Обновлена!"

class AminoacidsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            AminoAcidComposition.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('aminoacids_all')

def aminos_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = AminoAcidComposition.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("aminoacids_all")

        except AminoAcidComposition.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("aminoacids_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/aminoacids/aminoacids_all_delete_confirm.html", context)

def aminoacids_delete(request, id):
    context = {
            "is_active" : "aminoacids-panel",
            "active_aminoacids" : "active",
            "expand_aminoacids" : "show",
    }
    obj = get_object_or_404(AminoAcidComposition, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("aminoacids_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("aminoacids_delete")
 
    return render(request, "admin_templates/pages/aminoacids/aminoacids_delete_confirm.html", context)


# Chemical compositions CRUD operations

class ChemicalsView(View):
    model = Chemicals
    form_class = ChemicalsForm
    login_url = "login_page"
    active_panel = "chemicals-panel"
    success_url = reverse_lazy("chemicals_create")
    extra_context = {
        "is_active" : active_panel,
        "active_chemicals" : "active",
        "expand_chemicals" : "show",
        }
    
class ChemicalsListView(LoginRequiredMixin, ChemicalsView, ListView):
    template_name = "admin_templates/pages/chemicals/chemicals_list.html"
    paginate_by = 10

class ChemicalsCreateView(LoginRequiredMixin, SuccessMessageMixin, ChemicalsView, CreateView):
    template_name = 'admin_templates/pages/chemicals/chemicals_form.html'
    success_message = "Запись успешно Добавлена!" 

class ChemicalsUpdateView(LoginRequiredMixin, SuccessMessageMixin, ChemicalsView, UpdateView):
    template_name = "admin_templates/pages/chemicals/chemicals_form.html"
    success_message = "Запись успешно Обновлена!"

class ChemicalsDeleteView(View):
    def post(self, request):
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            Chemicals.objects.filter(id__in=selected_ids).delete()

            messages.success(request, 'Выбранные записи успешно удалены.')
        else:
            messages.error(request, 'Не выбраны записи для удаления.')

        return redirect('chemicals_all')

def chemicals_clear_clear(request):
    message = None
    if request.method =="POST":
        try:
            model = Chemicals.objects.all().delete()
            messages.success(request, "Все записи успешно удалены!")
            return redirect("chemicals_all")

        except Chemicals.DoesNotExist:
            messages.error(request, "Не удалось удалить все записи, повторите попытку!")
            return redirect("chemicals_delete")

    context={
        "message":message,
    }

    return render(request, "admin_templates/pages/chemicals/chemicals_all_delete_confirm.html", context)

def chemicals_delete(request, id):
    context = {
            "is_active" : "chemicals-panel",
            "active_chemicals" : "active",
            "expand_chemicals" : "show",
    }
    obj = get_object_or_404(Chemicals, id = id)
    if request.method =="POST":
        
        try:
            # delete object
            obj.delete()
            # after deleting redirect to
            # home page
            messages.success(request, "Запись успешно удалено!")
            return redirect("chemicals_all")
        except Exception as e:
            messages.error(request, "Не удалось удалить запись, повторите попытку!")
            return redirect("chemicals_delete")
    
    return render(request, "admin_templates/pages/chemicals/chemicals_delete_confirm.html", context)